import pandas as pd
import numpy as np
import logging
from openpyxl import load_workbook
from openpyxl.cell.rich_text import TextBlock, CellRichText

logger = logging.getLogger(__name__)

def extract_rich_text_from_cell(cell):
    """
    Extracts rich text formatting from an Excel cell.
    Returns a list of tuples: (text, is_subscript, is_superscript)
    """
    if cell.value is None:
        return []
    
    # Check if the cell contains rich text
    if isinstance(cell.value, CellRichText):
        result = []
        for text_block in cell.value:
            text = str(text_block)
            is_subscript = False
            is_superscript = False
            
            # Check if the text block has font formatting
            if hasattr(text_block, 'font') and text_block.font:
                if hasattr(text_block.font, 'vertAlign'):
                    if text_block.font.vertAlign == 'subscript':
                        is_subscript = True
                    elif text_block.font.vertAlign == 'superscript':
                        is_superscript = True
            
            result.append((text, is_subscript, is_superscript))
        return result
    else:
        # Plain text - return as a single block
        return [(str(cell.value), False, False)]

def read_excel_with_formatting(file_path, sheet_name=0):
    """
    Reads an Excel file and preserves rich text formatting (subscript/superscript).
    Returns a DataFrame with an additional column '_formatting' that contains formatting info.
    """
    # Load workbook with openpyxl to access formatting
    wb = load_workbook(file_path, data_only=False, rich_text=True)
    
    # Get the sheet
    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]
    
    # Read data with pandas for easy DataFrame creation
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    
    # Get header row (first row)
    headers = [cell.value for cell in ws[1]]
    
    # Create a dictionary to store formatting info for each cell
    # Format: {row_index: {column_name: [(text, is_subscript, is_superscript), ...]}}
    formatting_data = {}
    
    # Iterate through data rows (skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
        row_formatting = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                col_name = headers[col_idx]
                if col_name:  # Skip empty column names
                    rich_text = extract_rich_text_from_cell(cell)
                    if rich_text:
                        row_formatting[col_name] = rich_text
        
        if row_formatting:
            formatting_data[row_idx] = row_formatting
    
    # Add formatting data to DataFrame as a new column
    df['_formatting'] = df.index.map(lambda i: formatting_data.get(i, {}))
    
    wb.close()
    return df

def validate_excel_format(df):
    """
    Validates that a dataframe has the required columns and format
    """
    try:
        # Define required columns
        required_columns = ['Câu hỏi', 'A', 'B', 'C', 'D', 'đáp án']
        
        # Check if all required columns are present
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return {'error': f'Thiếu các cột yêu cầu: {", ".join(missing_columns)}'}
        
        # Check if any required columns are empty
        empty_columns = []
        for col in required_columns:
            # Check if the column has any NaN values
            if pd.isna(df[col]).any(skipna=False):
                empty_columns.append(col)
        
        if empty_columns:
            return {'error': f'Các cột sau đây chứa giá trị trống: {", ".join(empty_columns)}'}
        
        # Validate answer column values
        valid_answers = ['A', 'B', 'C', 'D']
        # Check for invalid answers
        invalid_answers = []
        for answer in df['đáp án'].unique():
            if answer not in valid_answers:
                invalid_answers.append(str(answer))
        
        if len(invalid_answers) > 0:
            return {'error': f'Tìm thấy giá trị đáp án không hợp lệ: {", ".join(invalid_answers)}. Đáp án hợp lệ là: A, B, C, D'}
        
        # Check if "Phân loại" column exists and has empty values
        if 'Phân loại' in df.columns:
            if pd.isna(df['Phân loại']).any():
                return {'error': 'Cột "Phân loại" chứa giá trị trống. Vui lòng điền đầy đủ phân loại cho tất cả câu hỏi'}
        
        return {'success': True}
    
    except Exception as e:
        logger.exception("Error validating Excel format")
        return {'error': f'Lỗi xử lý dữ liệu Excel: {str(e)}'}

def validate_excel_file(file_path):
    """
    Validates that the uploaded Excel file has at least one sheet with the required format
    """
    try:
        # Check if we can open the file as Excel
        try:
            excel_file = pd.ExcelFile(file_path)
        except Exception:
            return {'error': 'File không phải là file Excel hợp lệ'}
        
        # Get all sheet names
        sheet_names = excel_file.sheet_names
        if not sheet_names:
            return {'error': 'File Excel không chứa sheet nào'}
            
        # Validate first sheet to check format (use new function to preserve formatting)
        first_sheet = read_excel_with_formatting(file_path, sheet_name=sheet_names[0])
        return validate_excel_format(first_sheet)
    
    except Exception as e:
        logger.exception("Error validating Excel file")
        return {'error': f'Lỗi xử lý file Excel: {str(e)}'}

def get_random_questions(df, num_questions):
    """
    Gets a random sample of questions from the dataframe
    """
    # Check if we have enough questions
    total_questions = len(df)
    if total_questions < num_questions:
        raise ValueError(f"Not enough questions. Requested {num_questions}, but only {total_questions} are available.")
    
    # Get random sample
    sample_indices = np.random.choice(total_questions, num_questions, replace=False)
    sampled_questions = df.iloc[sample_indices].reset_index(drop=True)
    
    return sampled_questions
